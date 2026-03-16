#!/usr/bin/env python3
"""
MedTech Scout — FDA 510(k) Intelligence Extractor
===================================================
Pulls recent FDA 510(k) medical device clearances and enriches them
with LLM-powered classification to generate actionable sales intelligence
for MedTech service providers.

Built as a data source gap analysis demo for Zapyrus / Lumerate.

Usage:
    python medtech_scout.py              # Demo mode (sample data)
    python medtech_scout.py --live       # Live mode (pulls from openFDA API)
    python medtech_scout.py --days 60    # Live mode, last 60 days
    python medtech_scout.py --limit 50   # Live mode, up to 50 results

Requires:
    ANTHROPIC_API_KEY environment variable set
    pip install anthropic requests openpyxl
"""

import json
import os
import sys
import argparse
from datetime import datetime, timedelta

import anthropic
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────
# FDA 510(k) Data Fetching
# ──────────────────────────────────────────────

OPENFDA_510K_URL = "https://api.fda.gov/device/510k.json"

def fetch_live_clearances(days=30, limit=25):
    """Pull recent 510(k) clearances from openFDA API."""
    end = datetime.now()
    start = end - timedelta(days=days)
    date_range = f"[{start.strftime('%Y%m%d')} TO {end.strftime('%Y%m%d')}]"

    params = {
        "search": f"decision_date:{date_range} AND decision_code:SESE",
        "limit": limit,
        "sort": "decision_date:desc",
    }

    print(f"[*] Fetching 510(k) clearances from openFDA ({days} days, limit {limit})...")
    resp = requests.get(OPENFDA_510K_URL, params=params, timeout=30)
    resp.raise_for_status()
    results = resp.json().get("results", [])
    print(f"[+] Retrieved {len(results)} clearances")

    clearances = []
    for r in results:
        clearances.append({
            "k_number": r.get("k_number", ""),
            "device_name": r.get("device_name", ""),
            "applicant": r.get("applicant", ""),
            "decision_date": r.get("decision_date", ""),
            "clearance_type": r.get("clearance_type", ""),
            "product_code": r.get("product_code", ""),
            "advisory_committee_description": r.get("advisory_committee_description", ""),
            "statement_or_summary": r.get("statement_or_summary", ""),
            "expedited_review_flag": r.get("expedited_review_flag", ""),
            "third_party_flag": r.get("third_party_flag", ""),
        })
    return clearances


def load_sample_clearances():
    """Real FDA 510(k) clearances pulled from openFDA API (Feb 2026)."""
    return [
        {
            "k_number": "K254249",
            "device_name": "HKT Anatomical Locking Trauma System",
            "applicant": "Hankil Tech Medical Co., Ltd.",
            "decision_date": "2026-02-27",
            "clearance_type": "Traditional",
            "product_code": "HRS",
            "advisory_committee_description": "Orthopedic",
            "statement_or_summary": "Summary",
            "expedited_review_flag": "",
            "third_party_flag": "N",
        },
        {
            "k_number": "K253778",
            "device_name": "Allday Moisturizing Rinse",
            "applicant": "Elevate Oral Care",
            "decision_date": "2026-02-27",
            "clearance_type": "Traditional",
            "product_code": "LFD",
            "advisory_committee_description": "Dental",
            "statement_or_summary": "Summary",
            "expedited_review_flag": "",
            "third_party_flag": "N",
        },
        {
            "k_number": "K253797",
            "device_name": "One-Stop",
            "applicant": "Mediclus Co., Ltd.",
            "decision_date": "2026-02-27",
            "clearance_type": "Traditional",
            "product_code": "MVL",
            "advisory_committee_description": "Dental",
            "statement_or_summary": "Summary",
            "expedited_review_flag": "",
            "third_party_flag": "N",
        },
        {
            "k_number": "K253695",
            "device_name": "LigaMend",
            "applicant": "Riverpoint Medical, LLC",
            "decision_date": "2026-02-27",
            "clearance_type": "Traditional",
            "product_code": "MBI",
            "advisory_committee_description": "Orthopedic",
            "statement_or_summary": "Summary",
            "expedited_review_flag": "",
            "third_party_flag": "N",
        },
        {
            "k_number": "K253536",
            "device_name": "Evala Nerve Stimulator (EPNR002)",
            "applicant": "Epineuron Technologies, Inc.",
            "decision_date": "2026-02-27",
            "clearance_type": "Traditional",
            "product_code": "ETN",
            "advisory_committee_description": "Ear, Nose, Throat",
            "statement_or_summary": "Summary",
            "expedited_review_flag": "",
            "third_party_flag": "N",
        },
        {
            "k_number": "K252714",
            "device_name": "C-Lant Port",
            "applicant": "Vigor Medical Technologies, Ltd.",
            "decision_date": "2026-02-27",
            "clearance_type": "Traditional",
            "product_code": "GCJ",
            "advisory_committee_description": "Gastroenterology, Urology",
            "statement_or_summary": "Summary",
            "expedited_review_flag": "",
            "third_party_flag": "N",
        },
        {
            "k_number": "K253828",
            "device_name": "Medical Alexandrite and Nd:YAG Laser Therapy System (CM11LP)",
            "applicant": "Beijing HuaCheng Taike Technology Co., Ltd.",
            "decision_date": "2026-02-27",
            "clearance_type": "Traditional",
            "product_code": "GEX",
            "advisory_committee_description": "General, Plastic Surgery",
            "statement_or_summary": "Summary",
            "expedited_review_flag": "",
            "third_party_flag": "N",
        },
        {
            "k_number": "K251998",
            "device_name": "Atellica CH Diazo Total Bilirubin (D_TBil)",
            "applicant": "Siemens Healthcare Diagnostics, Inc.",
            "decision_date": "2026-02-27",
            "clearance_type": "Traditional",
            "product_code": "CIG",
            "advisory_committee_description": "Clinical Chemistry",
            "statement_or_summary": "Summary",
            "expedited_review_flag": "",
            "third_party_flag": "N",
        },
        {
            "k_number": "K253829",
            "device_name": "Medical Ultra-Pico Laser Treatment System (CM-SP-1064&532)",
            "applicant": "Beijing HuaCheng Taike Technology Co., Ltd.",
            "decision_date": "2026-02-27",
            "clearance_type": "Traditional",
            "product_code": "GEX",
            "advisory_committee_description": "General, Plastic Surgery",
            "statement_or_summary": "Summary",
            "expedited_review_flag": "",
            "third_party_flag": "N",
        },
        {
            "k_number": "K260292",
            "device_name": "HOTWIRE RF GUIDEWIRE",
            "applicant": "Atraverse Medical",
            "decision_date": "2026-02-27",
            "clearance_type": "Special",
            "product_code": "DXF",
            "advisory_committee_description": "Cardiovascular",
            "statement_or_summary": "Summary",
            "expedited_review_flag": "",
            "third_party_flag": "Y",
        },
        {
            "k_number": "K252448",
            "device_name": "AViTA Pulse Oximeter (SP61)",
            "applicant": "Avita Corporation",
            "decision_date": "2026-02-27",
            "clearance_type": "Traditional",
            "product_code": "DQA",
            "advisory_committee_description": "Cardiovascular",
            "statement_or_summary": "Summary",
            "expedited_review_flag": "",
            "third_party_flag": "N",
        },
        {
            "k_number": "K253535",
            "device_name": "Ligence Heart",
            "applicant": "Ligence UAB",
            "decision_date": "2026-02-27",
            "clearance_type": "Traditional",
            "product_code": "QIH",
            "advisory_committee_description": "Radiology",
            "statement_or_summary": "Statement",
            "expedited_review_flag": "",
            "third_party_flag": "N",
        },
        {
            "k_number": "K254017",
            "device_name": "SWINGO-3D Lumbar Cage System",
            "applicant": "Implanet",
            "decision_date": "2026-02-26",
            "clearance_type": "Traditional",
            "product_code": "MAX",
            "advisory_committee_description": "Orthopedic",
            "statement_or_summary": "Summary",
            "expedited_review_flag": "",
            "third_party_flag": "N",
        },
        {
            "k_number": "K251700",
            "device_name": "Ganymede Wound Care Matrix",
            "applicant": "Speciality Fibres and Materials Limited",
            "decision_date": "2026-02-26",
            "clearance_type": "Traditional",
            "product_code": "QSY",
            "advisory_committee_description": "General, Plastic Surgery",
            "statement_or_summary": "Summary",
            "expedited_review_flag": "",
            "third_party_flag": "N",
        },
        {
            "k_number": "K254251",
            "device_name": "InnovexView Hysteroscope",
            "applicant": "Shanghai AnQing Medical Instrument Co., Ltd.",
            "decision_date": "2026-02-26",
            "clearance_type": "Traditional",
            "product_code": "HIH",
            "advisory_committee_description": "Obstetrics/Gynecology",
            "statement_or_summary": "Summary",
            "expedited_review_flag": "",
            "third_party_flag": "N",
        },
    ]


# ──────────────────────────────────────────────
# LLM Enrichment Pipeline
# ──────────────────────────────────────────────

ENRICHMENT_SYSTEM_PROMPT = """You are a MedTech market intelligence analyst. Your job is to classify
FDA 510(k) device clearances and extract structured sales intelligence that would help
MedTech service providers (CROs, CDMOs, quality/regulatory consultants, contract design
and engineering firms) identify business opportunities.

For each device clearance, you must return a JSON array of objects with these exact fields:

- "k_number": the 510(k) number (pass through from input)
- "therapeutic_area": one of [Cardiovascular, Neurology, Orthopedic, Oncology, Respiratory,
  Gastroenterology/Urology, Ophthalmology, Dermatology, Diagnostics/Imaging, General Surgery,
  Renal, Dental, ENT, Wound Care, Other]
- "device_category": a short classification (e.g. "Surgical Robot", "Wearable Monitor",
  "Implantable Device", "Diagnostic Imaging", "In Vitro Diagnostic", "Therapeutic Device",
  "Software as Medical Device", "Catheter/Delivery System", "Prosthetic/Implant")
- "technology_keywords": array of 3-5 technology keywords relevant to the device
- "commercial_stage": one of ["Early Commercial", "Growth Phase", "Established/Iterative"]
  based on whether this appears to be a new entrant, a next-gen product, or incremental update
- "service_opportunities": array of 2-4 MedTech services this company likely needs
  (e.g. "Contract Manufacturing", "Regulatory Consulting", "Clinical Trial Management",
  "Quality System Implementation", "Sterilization Services", "Biocompatibility Testing",
  "Design Verification & Validation", "Post-Market Surveillance", "Packaging & Labeling")
- "sales_trigger_summary": a 1-2 sentence actionable insight a sales rep could use to
  reach out to this company. Be specific about WHY this clearance matters.
- "urgency_score": 1-5 integer rating of how time-sensitive this opportunity is for a
  service provider (5 = act now, 1 = long-term pipeline)

Return ONLY valid JSON. No markdown, no explanation. Just the JSON array."""


def enrich_batch(client, clearances, batch_size=5):
    """Send clearances to Claude in batches for classification."""
    all_enriched = []

    for i in range(0, len(clearances), batch_size):
        batch = clearances[i : i + batch_size]
        batch_num = (i // batch_size) + 1
        total_batches = (len(clearances) + batch_size - 1) // batch_size
        print(f"[*] Enriching batch {batch_num}/{total_batches} ({len(batch)} clearances)...")

        input_data = json.dumps(batch, indent=2)

        try:
            response = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=4096,
                system=ENRICHMENT_SYSTEM_PROMPT,
                messages=[
                    {
                        "role": "user",
                        "content": f"Classify and enrich these FDA 510(k) clearances:\n\n{input_data}",
                    }
                ],
            )

            raw = response.content[0].text.strip()
            # Strip markdown fences if present
            if raw.startswith("```"):
                raw = raw.split("\n", 1)[1]
                raw = raw.rsplit("```", 1)[0]
            parsed = json.loads(raw)
            all_enriched.extend(parsed)
            print(f"[+] Batch {batch_num} enriched successfully")

        except json.JSONDecodeError as e:
            print(f"[!] JSON parse error on batch {batch_num}: {e}")
            print(f"    Raw response: {raw[:200]}...")
            # Fall back: create unenriched entries
            for c in batch:
                all_enriched.append({
                    "k_number": c["k_number"],
                    "therapeutic_area": c.get("advisory_committee_description", "Unknown"),
                    "device_category": "Unclassified",
                    "technology_keywords": [],
                    "commercial_stage": "Unknown",
                    "service_opportunities": [],
                    "sales_trigger_summary": "LLM classification failed — manual review needed.",
                    "urgency_score": 0,
                })
        except Exception as e:
            print(f"[!] API error on batch {batch_num}: {e}")
            for c in batch:
                all_enriched.append({
                    "k_number": c["k_number"],
                    "therapeutic_area": "Error",
                    "device_category": "Error",
                    "technology_keywords": [],
                    "commercial_stage": "Error",
                    "service_opportunities": [],
                    "sales_trigger_summary": f"API error: {e}",
                    "urgency_score": 0,
                })

    return all_enriched


def demo_enrichment(clearances):
    """Pre-generated enrichment for real FDA 510(k) clearances (Feb 2026)."""
    demo_data = {
        "K254249": {
            "therapeutic_area": "Orthopedic",
            "device_category": "Prosthetic/Implant",
            "technology_keywords": ["trauma fixation", "locking plate", "bone fracture", "anatomical plating", "titanium implant"],
            "commercial_stage": "Growth Phase",
            "service_opportunities": ["Contract Manufacturing", "Sterilization Services", "Biocompatibility Testing", "Regulatory Consulting"],
            "sales_trigger_summary": "Hankil Tech (South Korea) clearing a trauma plating system in the US signals international market expansion. They'll need US-based contract manufacturing, sterilization validation, and distribution partners to scale domestically.",
            "urgency_score": 4,
        },
        "K253778": {
            "therapeutic_area": "Dental",
            "device_category": "Therapeutic Device",
            "technology_keywords": ["oral care", "moisturizing rinse", "xerostomia", "dental hygiene", "mucosal care"],
            "commercial_stage": "Established/Iterative",
            "service_opportunities": ["Contract Manufacturing", "Packaging & Labeling", "Regulatory Consulting", "Post-Market Surveillance"],
            "sales_trigger_summary": "Elevate Oral Care is expanding their dental product line with a new moisturizing rinse formulation. As a specialty oral care company, they likely outsource manufacturing — packaging and contract fill partners should engage.",
            "urgency_score": 2,
        },
        "K253797": {
            "therapeutic_area": "Dental",
            "device_category": "Implantable Device",
            "technology_keywords": ["dental implant", "one-piece implant", "osseointegration", "titanium dental", "oral surgery"],
            "commercial_stage": "Growth Phase",
            "service_opportunities": ["Contract Manufacturing", "Biocompatibility Testing", "Sterilization Services", "Quality System Implementation"],
            "sales_trigger_summary": "Mediclus (South Korea) is entering the US dental implant market — a highly competitive space requiring precision manufacturing and rigorous biocompatibility testing. US-based manufacturing and quality consultants should reach out now.",
            "urgency_score": 4,
        },
        "K253695": {
            "therapeutic_area": "Orthopedic",
            "device_category": "Implantable Device",
            "technology_keywords": ["ligament repair", "meniscal fixation", "arthroscopic surgery", "soft tissue repair", "suture anchor"],
            "commercial_stage": "Growth Phase",
            "service_opportunities": ["Contract Manufacturing", "Sterilization Services", "Clinical Trial Management", "Design Verification & Validation"],
            "sales_trigger_summary": "Riverpoint Medical's LigaMend clearance for soft tissue repair positions them in the growing arthroscopic surgery market. Post-clearance manufacturing scale-up and V&V for design iterations are immediate needs.",
            "urgency_score": 4,
        },
        "K253536": {
            "therapeutic_area": "ENT",
            "device_category": "Therapeutic Device",
            "technology_keywords": ["nerve stimulation", "peripheral nerve repair", "electrical stimulation", "neuromuscular", "nerve regeneration"],
            "commercial_stage": "Early Commercial",
            "service_opportunities": ["Contract Manufacturing", "Regulatory Consulting", "Clinical Trial Management", "Quality System Implementation"],
            "sales_trigger_summary": "Epineuron Technologies is a startup-stage company with a novel nerve stimulation platform. As an early-commercial company they'll need full-service support — manufacturing, clinical evidence generation, and quality system buildout.",
            "urgency_score": 5,
        },
        "K252714": {
            "therapeutic_area": "Gastroenterology/Urology",
            "device_category": "Implantable Device",
            "technology_keywords": ["implantable port", "vascular access", "infusion port", "central venous access", "chemotherapy delivery"],
            "commercial_stage": "Growth Phase",
            "service_opportunities": ["Contract Manufacturing", "Sterilization Services", "Biocompatibility Testing", "Packaging & Labeling"],
            "sales_trigger_summary": "Vigor Medical Technologies (international) clearing an implantable port for the US market indicates commercial expansion. Implantable ports require strict sterilization and biocompatibility protocols — specialized service providers should engage.",
            "urgency_score": 3,
        },
        "K253828": {
            "therapeutic_area": "Dermatology",
            "device_category": "Therapeutic Device",
            "technology_keywords": ["alexandrite laser", "Nd:YAG laser", "aesthetic dermatology", "laser therapy", "skin treatment"],
            "commercial_stage": "Growth Phase",
            "service_opportunities": ["Regulatory Consulting", "Quality System Implementation", "Design Verification & Validation", "Post-Market Surveillance"],
            "sales_trigger_summary": "Beijing HuaCheng is pushing dual-wavelength laser systems into the US aesthetic market. Chinese manufacturers entering the US need extensive regulatory and quality system support for FDA compliance — consultants should act quickly.",
            "urgency_score": 4,
        },
        "K251998": {
            "therapeutic_area": "Diagnostics/Imaging",
            "device_category": "In Vitro Diagnostic",
            "technology_keywords": ["bilirubin assay", "clinical chemistry", "automated analyzer", "IVD reagent", "hepatic function"],
            "commercial_stage": "Established/Iterative",
            "service_opportunities": ["Post-Market Surveillance", "Quality System Implementation", "Regulatory Consulting", "Packaging & Labeling"],
            "sales_trigger_summary": "Siemens Healthcare Diagnostics expanding their Atellica chemistry menu is routine iteration for an established IVD platform. Low urgency but steady opportunity for reagent manufacturing and compliance support partners.",
            "urgency_score": 1,
        },
        "K253829": {
            "therapeutic_area": "Dermatology",
            "device_category": "Therapeutic Device",
            "technology_keywords": ["picosecond laser", "tattoo removal", "pigment treatment", "aesthetic laser", "skin rejuvenation"],
            "commercial_stage": "Growth Phase",
            "service_opportunities": ["Regulatory Consulting", "Quality System Implementation", "Design Verification & Validation", "Post-Market Surveillance"],
            "sales_trigger_summary": "Second laser system from Beijing HuaCheng cleared the same day — they're aggressively expanding their US aesthetic device portfolio. Dual clearances signal a serious market push; regulatory and quality partners are needed immediately.",
            "urgency_score": 4,
        },
        "K260292": {
            "therapeutic_area": "Cardiovascular",
            "device_category": "Catheter/Delivery System",
            "technology_keywords": ["RF guidewire", "chronic total occlusion", "CTO crossing", "radiofrequency", "interventional cardiology"],
            "commercial_stage": "Early Commercial",
            "service_opportunities": ["Contract Manufacturing", "Sterilization Services", "Clinical Trial Management", "Regulatory Consulting"],
            "sales_trigger_summary": "Special 510(k) clearance with third-party review for an RF-enabled guidewire — Atraverse Medical is a newer entrant in the CTO crossing space. This is a high-acuity, high-value device category with complex manufacturing needs. Engage now.",
            "urgency_score": 5,
        },
        "K252448": {
            "therapeutic_area": "Cardiovascular",
            "device_category": "Wearable Monitor",
            "technology_keywords": ["pulse oximetry", "SpO2 monitoring", "vital signs", "patient monitoring", "portable diagnostics"],
            "commercial_stage": "Growth Phase",
            "service_opportunities": ["Contract Manufacturing", "Design Verification & Validation", "Packaging & Labeling", "Post-Market Surveillance"],
            "sales_trigger_summary": "Avita Corporation clearing a new pulse oximeter model suggests product line expansion. Pulse ox is a crowded market — they'll need competitive manufacturing costs and efficient packaging to win market share.",
            "urgency_score": 2,
        },
        "K253535": {
            "therapeutic_area": "Diagnostics/Imaging",
            "device_category": "Software as Medical Device",
            "technology_keywords": ["AI echocardiography", "cardiac imaging AI", "automated echo analysis", "machine learning", "cardiac function"],
            "commercial_stage": "Early Commercial",
            "service_opportunities": ["Regulatory Consulting", "Clinical Trial Management", "Quality System Implementation", "Post-Market Surveillance"],
            "sales_trigger_summary": "Ligence (Lithuania) clearing an AI-powered cardiac imaging platform in the US is a major market entry signal. SaMD companies need specialized regulatory, clinical evidence, and algorithm change management support — high-value engagement.",
            "urgency_score": 5,
        },
        "K254017": {
            "therapeutic_area": "Orthopedic",
            "device_category": "Prosthetic/Implant",
            "technology_keywords": ["lumbar fusion", "interbody cage", "3D-printed implant", "spinal surgery", "vertebral stabilization"],
            "commercial_stage": "Growth Phase",
            "service_opportunities": ["Contract Manufacturing", "Biocompatibility Testing", "Sterilization Services", "Design Verification & Validation"],
            "sales_trigger_summary": "Implanet (France) expanding their spinal implant portfolio into the US market with a 3D lumbar cage system. International companies scaling in the US need domestic manufacturing and testing partners for supply chain efficiency.",
            "urgency_score": 3,
        },
        "K251700": {
            "therapeutic_area": "Wound Care",
            "device_category": "Therapeutic Device",
            "technology_keywords": ["wound dressing", "wound care matrix", "advanced wound management", "bioactive fibers", "tissue regeneration"],
            "commercial_stage": "Early Commercial",
            "service_opportunities": ["Contract Manufacturing", "Sterilization Services", "Biocompatibility Testing", "Regulatory Consulting"],
            "sales_trigger_summary": "Speciality Fibres and Materials (UK) clearing a full wound care matrix product line (multiple sizes) signals aggressive US market entry. They'll need US manufacturing, sterilization, and distribution infrastructure — act now before they lock in partners.",
            "urgency_score": 5,
        },
        "K254251": {
            "therapeutic_area": "Obstetrics/Gynecology",
            "device_category": "Diagnostic Imaging",
            "technology_keywords": ["hysteroscopy", "endoscopic visualization", "gynecological imaging", "minimally invasive", "uterine diagnosis"],
            "commercial_stage": "Growth Phase",
            "service_opportunities": ["Contract Manufacturing", "Sterilization Services", "Design Verification & Validation", "Packaging & Labeling"],
            "sales_trigger_summary": "Shanghai AnQing is entering the US hysteroscopy market with multiple scope configurations. Chinese endoscope manufacturers entering the US need quality system alignment and manufacturing partners to compete with established players like Olympus and Karl Storz.",
            "urgency_score": 3,
        },
    }

    enriched = []
    for c in clearances:
        k = c["k_number"]
        if k in demo_data:
            entry = {"k_number": k, **demo_data[k]}
        else:
            entry = {
                "k_number": k,
                "therapeutic_area": c.get("advisory_committee_description", "Unknown"),
                "device_category": "Unclassified",
                "technology_keywords": [],
                "commercial_stage": "Unknown",
                "service_opportunities": [],
                "sales_trigger_summary": "No demo data available — run with ANTHROPIC_API_KEY for live classification.",
                "urgency_score": 0,
            }
        enriched.append(entry)
    return enriched


# ──────────────────────────────────────────────
# Spreadsheet Output
# ──────────────────────────────────────────────

# Color palette (inspired by Zapyrus brand)
HEADER_FILL = PatternFill("solid", fgColor="1B2A4A")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="E8EDF4")
SUBHEADER_FONT = Font(name="Arial", bold=True, color="1B2A4A", size=10)
DATA_FONT = Font(name="Arial", size=10, color="333333")
URGENCY_COLORS = {
    5: PatternFill("solid", fgColor="FEE2E2"),  # Red - Act now
    4: PatternFill("solid", fgColor="FEF3C7"),  # Amber
    3: PatternFill("solid", fgColor="FEF9C3"),  # Yellow
    2: PatternFill("solid", fgColor="ECFDF5"),  # Light green
    1: PatternFill("solid", fgColor="F0F9FF"),  # Light blue
}
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def build_spreadsheet(clearances, enriched, output_path):
    """Build a polished xlsx with raw data + enriched intelligence."""
    wb = Workbook()

    # ── Sheet 1: Enriched Intelligence (the star of the show) ──
    ws = wb.active
    ws.title = "Sales Intelligence"
    ws.sheet_properties.tabColor = "1B2A4A"

    headers = [
        ("510(k) #", 12),
        ("Device Name", 35),
        ("Company", 30),
        ("Decision Date", 14),
        ("Therapeutic Area", 20),
        ("Device Category", 22),
        ("Technology Keywords", 32),
        ("Commercial Stage", 18),
        ("Service Opportunities", 38),
        ("Sales Trigger Summary", 50),
        ("Urgency", 10),
    ]

    # Title row
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = "MedTech Scout — FDA 510(k) Sales Intelligence Report"
    title_cell.font = Font(name="Arial", bold=True, size=14, color="1B2A4A")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtitle
    ws.merge_cells("A2:K2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} | {len(clearances)} clearances analyzed | Data source: openFDA 510(k) API"
    sub_cell.font = Font(name="Arial", size=9, italic=True, color="6B7280")
    ws.row_dimensions[2].height = 20

    # Headers (row 4)
    header_row = 4
    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 25

    # Map enriched data by k_number for lookup
    enriched_map = {e["k_number"]: e for e in enriched}

    # Data rows
    for row_idx, c in enumerate(clearances, header_row + 1):
        e = enriched_map.get(c["k_number"], {})
        urgency = e.get("urgency_score", 0)

        row_data = [
            c["k_number"],
            c["device_name"],
            c["applicant"],
            c["decision_date"],
            e.get("therapeutic_area", ""),
            e.get("device_category", ""),
            ", ".join(e.get("technology_keywords", [])),
            e.get("commercial_stage", ""),
            ", ".join(e.get("service_opportunities", [])),
            e.get("sales_trigger_summary", ""),
            urgency,
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Color urgency column
            if col_idx == 11 and urgency in URGENCY_COLORS:
                cell.fill = URGENCY_COLORS[urgency]
                cell.alignment = Alignment(horizontal="center", vertical="top")

        ws.row_dimensions[row_idx].height = 55

    # Freeze header
    ws.freeze_panes = "A5"

    # AutoFilter
    ws.auto_filter.ref = f"A{header_row}:K{header_row + len(clearances)}"

    # ── Sheet 2: Data Source Gap Analysis ──
    ws2 = wb.create_sheet("Data Source Gaps")
    ws2.sheet_properties.tabColor = "10B981"

    ws2.merge_cells("A1:E1")
    ws2["A1"].value = "Zapyrus Data Source Gap Analysis"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1B2A4A")
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells("A2:E2")
    ws2["A2"].value = "Potential new data sources identified through analysis of Zapyrus product features vs. publicly available MedTech data"
    ws2["A2"].font = Font(name="Arial", size=9, italic=True, color="6B7280")

    gap_headers = [
        ("Data Source", 32),
        ("Type", 18),
        ("Signal Value for Sales Teams", 45),
        ("Coverage Gap", 35),
        ("Difficulty to Integrate", 20),
    ]

    for col_idx, (h, w) in enumerate(gap_headers, 1):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws2.column_dimensions[get_column_letter(col_idx)].width = w

    gaps = [
        [
            "EUDAMED (EU Medical Device Database)",
            "Regulatory",
            "EU device registrations, UDI data, notified body certificates, and vigilance reports — critical for companies selling into EU markets",
            "Zapyrus appears heavily US/FDA-focused; EU regulatory signals are a major blind spot for globally-selling customers",
            "Medium",
        ],
        [
            "USPTO / EPO / WIPO Patent Filings",
            "IP / R&D",
            "Patent activity is a 12-24 month leading indicator — a company filing device patents is in active development before clinical trials begin",
            "Zapyrus tracks clinical milestones but not upstream IP activity that signals earlier-stage opportunities",
            "Medium",
        ],
        [
            "SEC EDGAR / SEDAR+ Filings",
            "Financial / Strategic",
            "10-K, S-1, and prospectus filings contain forward-looking statements on device pipeline timelines, manufacturing partnerships, and regulatory submission plans",
            "Funding announcements are tracked, but the raw filings contain far richer detail on strategic direction",
            "Low-Medium",
        ],
        [
            "WHO ICTRP (International Clinical Trials Registry)",
            "Clinical",
            "Aggregates non-US registries: CTRI (India), ANZCTR (Australia/NZ), ChiCTR (China), EU CTR — surfaces global trial activity",
            "ClinicalTrials.gov is likely the primary source; international registries expand coverage significantly for global MedTech",
            "Low",
        ],
        [
            "Government Procurement (SAM.gov, MERX, TED)",
            "Procurement / Intent",
            "MedTech companies winning government contracts or responding to tenders signals immediate purchasing intent and confirmed budget",
            "Not currently visible in Zapyrus signal types; procurement data is a direct buying signal",
            "Medium",
        ],
        [
            "Accelerator / Incubator Cohorts",
            "Company Discovery",
            "Programs like MedTech Innovator, JLABS, Plug and Play Health, Y Combinator announce cohorts publicly — surfaces stealth-mode companies months before press coverage",
            "Zapyrus tracks 'new companies' but cohort announcements are a systematic early-detection source",
            "Low",
        ],
        [
            "FDA 510(k) / PMA / De Novo Raw Databases",
            "Regulatory / Predictive",
            "Submission dates, predicate devices, and review timelines feed predictive models for commercial readiness — this demo is an example",
            "Device approvals are tracked, but raw submission data enables earlier and more granular signals",
            "Low",
        ],
        [
            "Job Posting Data (LinkedIn, Indeed, Careers Pages)",
            "Intent / Lifecycle",
            "WHAT a MedTech company is hiring for (regulatory, clinical ops, manufacturing scale-up) signals where they are in the product lifecycle",
            "Job changes are tracked for contacts; job POSTINGS as a company lifecycle signal appear untapped",
            "Medium-High",
        ],
    ]

    for row_idx, gap in enumerate(gaps, 5):
        for col_idx, val in enumerate(gap, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws2.row_dimensions[row_idx].height = 65

    ws2.freeze_panes = "A5"

    # ── Sheet 3: Raw FDA Data ──
    ws3 = wb.create_sheet("Raw 510(k) Data")
    ws3.sheet_properties.tabColor = "6B7280"

    raw_headers = [
        ("510(k) #", 12),
        ("Device Name", 35),
        ("Applicant", 30),
        ("Decision Date", 14),
        ("Clearance Type", 16),
        ("Product Code", 14),
        ("Advisory Committee", 25),
        ("Summary/Statement", 16),
        ("Expedited Review", 14),
        ("Third Party", 12),
    ]

    for col_idx, (h, w) in enumerate(raw_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER
        ws3.column_dimensions[get_column_letter(col_idx)].width = w

    for row_idx, c in enumerate(clearances, 2):
        vals = [
            c["k_number"],
            c["device_name"],
            c["applicant"],
            c["decision_date"],
            c["clearance_type"],
            c["product_code"],
            c["advisory_committee_description"],
            c["statement_or_summary"],
            c["expedited_review_flag"],
            c["third_party_flag"],
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

    ws3.freeze_panes = "A2"

    # Save
    wb.save(output_path)
    print(f"[+] Spreadsheet saved: {output_path}")


# ──────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="MedTech Scout — FDA 510(k) Intelligence Extractor")
    parser.add_argument("--live", action="store_true", help="Pull live data from openFDA API")
    parser.add_argument("--days", type=int, default=30, help="Days of history to fetch (default: 30)")
    parser.add_argument("--limit", type=int, default=25, help="Max clearances to fetch (default: 25)")
    parser.add_argument("--output", type=str, default=None, help="Output file path")
    args = parser.parse_args()

    # Check API key
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")

    # Fetch data
    if args.live:
        clearances = fetch_live_clearances(days=args.days, limit=args.limit)
        if not clearances:
            print("[!] No clearances returned. Try increasing --days or --limit.")
            sys.exit(1)
    else:
        print("[*] Running in demo mode (sample data). Use --live for real FDA data.")
        clearances = load_sample_clearances()

    # Enrich with LLM (or demo fallback)
    if api_key:
        client = anthropic.Anthropic(api_key=api_key)
        enriched = enrich_batch(client, clearances, batch_size=5)
    else:
        print("[*] No ANTHROPIC_API_KEY found — using demo enrichment.")
        print("    Set your key to enable live LLM classification:")
        print('    export ANTHROPIC_API_KEY="sk-ant-..."')
        enriched = demo_enrichment(clearances)

    # Build output
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = args.output or f"medtech_scout_report_{timestamp}.xlsx"
    build_spreadsheet(clearances, enriched, output_path)

    print(f"\n{'='*60}")
    print(f"  MedTech Scout Report Complete")
    print(f"  {len(clearances)} clearances fetched")
    print(f"  {len(enriched)} enriched with LLM classification")
    print(f"  Output: {output_path}")
    print(f"{'='*60}")

    return output_path


if __name__ == "__main__":
    main()
